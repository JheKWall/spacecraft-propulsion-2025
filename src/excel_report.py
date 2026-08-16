"""Build the Excel workbook.

Five sheets. The one that matters is Measurements: every figure with its
citation, page, and the verbatim sentence it came from. That sheet is what makes
the analysis defensible - any number on any chart can be traced from it to a
source in a few seconds, without opening the database.
"""
import mysql.connector
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import DB, OUTPUT_DIR, PROCESSED_DIR, db_password
from analyze import build_analysis, fetch, query

HEADER_FILL = PatternFill("solid", fgColor="2A78D6")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

# Column widths per sheet: (column name, width). Anything unlisted gets a default.
WIDTHS = {
    "citation": 62, "quoted_text": 80, "note": 70, "name": 24, "subtype": 26,
    "role": 22, "authors": 30, "publisher": 30, "url": 46, "local_file": 34,
    "propellant": 16, "finding": 30, "detail": 96,
}


def _write_sheet(writer, df, sheet_name, wrap_columns=()):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    for i, column in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(column, 16)

    if len(df):
        ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for column in wrap_columns:
        if column not in df.columns:
            continue
        letter = get_column_letter(list(df.columns).index(column) + 1)
        for row in range(2, len(df) + 2):
            ws[f"{letter}{row}"].alignment = Alignment(wrap_text=True, vertical="top")


def findings_frame(analysis):
    computable = analysis[analysis["tw_low"].notna()]
    theoretical_top = analysis[analysis["maturity"] == "theoretical"]["isp_high"].max()
    proven_top = analysis[analysis["maturity"] == "proven"]["isp_high"].max()
    launch = computable[computable["launch_capable"]]

    rows = [
        ("Research question",
         "Which mission role is each category of spacecraft propulsion best suited "
         "for? Compared on specific impulse, thrust, input power, and "
         "thrust-to-weight ratio."),
        ("Chemical - verdict",
         "The only category that can launch from a planetary surface. But not all of "
         "it: the MR-103J is a flight-proven chemical thruster with a T/W of "
         "0.05-0.33 and cannot lift itself. Best suited to launch and high-thrust "
         "roles, which nothing else can perform at all."),
        ("Electric - verdict",
         "Every electric thruster in this dataset beats every chemical engine on "
         "specific impulse, and none produces more than a quarter of a newton. "
         "Thrust-to-weight is of order 0.0002-0.002 against roughly 60 for the "
         "RS-25. Best suited to station-keeping and long-duration deep-space "
         "roles where time is available and thrust is not required."),
        ("Nuclear - verdict",
         "Occupies a high-Isp, high-thrust middle that would suit upper stages and "
         "deep-space primaries. No nuclear system in this dataset has ever flown. "
         "DRACO, the only one being built to fly, was cancelled in 2025 and "
         "published no performance figure at any tier."),
        ("Specific impulse alone misleads",
         f"The two highest-Isp systems have never been built. The theoretical "
         f"maximum ({theoretical_top:,.0f} s) is {theoretical_top / proven_top:,.0f} "
         f"times the best figure for any hardware ever fired ({proven_top:,.0f} s). "
         f"The original class project ranked on this metric alone."),
        ("Thrust-to-weight coverage",
         f"Computable for only {len(computable)} of {len(analysis)} systems. "
         f"{int(analysis['thrust_low'].notna().sum())} systems have a citable "
         f"thrust; only {int(analysis['engine_mass_low'].notna().sum())} have a "
         f"citable engine mass. The metric that answers 'can this leave the ground' "
         f"is the one the literature is least willing to supply."),
        ("Systems that can leave Earth",
         f"{len(launch)} of {len(computable)} with a computable T/W exceed 1.0: "
         + ", ".join(launch.sort_values("tw_high", ascending=False)["name"])),
        ("Flight heritage",
         f"{int((~analysis['flown']).sum())} of {len(analysis)} systems have never "
         f"flown, and every one of them is nuclear."),
        ("Data-quality finding",
         "Frisbee (2003) Table 3 contains a unit error: its km/s column is 1000x too "
         "large for the three exotic rows, giving antimatter an exhaust velocity 333 "
         "times the speed of light. The paper's own body text gives the correct "
         "figure. The original class project transcribed the table faithfully; the "
         "error is upstream. See docs/data-collection.md section 3."),
        ("Source conflicts",
         "Recorded rather than resolved by preference. RS-25 dry mass differs 10% "
         "between NASA and secondary sources, moving its T/W from ~73 to 61-66. The "
         "BHT-600 has two citable and different power ranges, stored as two rows. "
         "See docs/data-collection.md section 5."),
        ("Cost",
         "Excluded. The original project statement asked for the most cost-effective "
         "system. No per-system cost figure exists in any consulted source, and none "
         "exists publicly for the theoretical concepts. The question was changed "
         "rather than answered badly."),
        ("Verification",
         "Every figure carries a citation, page, and verbatim quote (see the "
         "Measurements sheet). 5 of 40 figures were independently re-extracted by a "
         "different model family reading cold; all 5 confirmed. The remaining 35 are "
         "hand-check-only and marked as such in output/verification_checklist.md."),
    ]
    return pd.DataFrame(rows, columns=["finding", "detail"])


def main():
    conn = mysql.connector.connect(**DB, password=db_password())
    measurements = fetch(conn)
    systems = query(conn, "SELECT * FROM propulsion_system ORDER BY system_id")
    sources = query(conn, "SELECT * FROM source ORDER BY source_id")
    conn.close()

    analysis = build_analysis(measurements, systems)

    comparison = analysis[[
        "name", "category", "subtype", "role", "maturity", "flown",
        "isp_low", "isp_high", "thrust_low", "thrust_high",
        "engine_mass_low", "engine_mass_high", "input_power_low",
        "input_power_high", "tw_low", "tw_high", "launch_capable",
    ]]

    target = OUTPUT_DIR / "Spacecraft_Propulsion_Analysis.xlsx"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        _write_sheet(writer, findings_frame(analysis), "Findings",
                     wrap_columns=("detail",))
        _write_sheet(writer, comparison, "Comparison")
        _write_sheet(writer, measurements, "Measurements",
                     wrap_columns=("quoted_text", "note", "citation"))
        _write_sheet(writer, systems.drop(columns=["system_id"]), "Systems")
        _write_sheet(writer, sources.drop(columns=["source_id"]), "Sources",
                     wrap_columns=("citation",))

    analysis.to_csv(PROCESSED_DIR / "analysis.csv", index=False)
    print(f"  wrote {target.name}")
    print(f"    Findings {len(findings_frame(analysis))} rows | "
          f"Comparison {len(comparison)} | Measurements {len(measurements)} | "
          f"Systems {len(systems)} | Sources {len(sources)}")


if __name__ == "__main__":
    main()
